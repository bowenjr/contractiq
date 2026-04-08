"""
KnowledgeIO — export / import knowledge base tables to/from Excel.
"""

from datetime import date
from pathlib import Path


# Map table names → (get_all method name, create method name)
# Handles cases where DB method doesn't follow get_all_{table} / create_{table} exactly.
_TABLE_METHOD_MAP = {
    "company_positions":   ("get_all_company_positions",    "create_company_position"),
    "insurance_positions": ("get_all_insurance_positions",  "create_insurance_position"),
    "escalation_rules":    ("get_all_escalation_rules",     "create_escalation_rule"),
    "product_risk_profiles": ("get_all_product_risk_profiles", "create_product_risk_profile"),
    "commercial_term_library": ("get_all_commercial_terms",  "create_commercial_term"),
    "product_term_risk_map": ("get_all_product_term_maps",  "create_product_term_map"),
    "deliverable_templates": ("get_all_deliverable_templates", "create_deliverable_template"),
    "clause_playbooks":    ("get_all_clause_playbooks",     "create_clause_playbook"),
    "review_routing_rules": ("get_all_routing_rules",       "create_routing_rule"),
    "negotiation_history": ("get_all_negotiation_history",  "create_negotiation_record"),
    "supplier_intelligence": ("get_all_supplier_intelligence", "create_supplier_intel"),
    "project_type_profiles": ("get_all_project_type_profiles", "create_project_type_profile"),
    "jurisdiction_rules":  ("get_all_jurisdiction_rules",   "create_jurisdiction_rule"),
}

SUPPORTED_TABLES = list(_TABLE_METHOD_MAP.keys())


class KnowledgeIO:

    def __init__(self, db):
        self.db = db

    # ── Export single table ───────────────────────────────────────────────────

    def export_table_to_excel(self, table_name: str, output_path) -> bool:
        """Export a single reference table to Excel. Returns True on success."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        records = self._get_all_records(table_name)
        if not records:
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]

        header_fill = PatternFill(fill_type="solid", fgColor="1a1f2e")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        inactive_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

        headers = list(records[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for row_idx, record in enumerate(records, 2):
            for col_idx, value in enumerate(record.values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                if record.get("active") == 0:
                    cell.fill = inactive_fill

        # Auto-fit columns (capped)
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(
                max(max_len + 2, 12), 60
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(str(output_path))
        return True

    # ── Import single table ───────────────────────────────────────────────────

    def import_table_from_excel(self, table_name: str, file_path) -> dict:
        """
        Import records from an Excel file into a reference table.
        id column is stripped so DB assigns new IDs.
        Returns: {imported, skipped, errors}
        """
        import openpyxl

        results = {"imported": 0, "skipped": 0, "errors": []}

        _, create_method_name = _TABLE_METHOD_MAP.get(table_name, (None, None))
        if not create_method_name:
            results["errors"].append(f"Table '{table_name}' not supported")
            return results

        create_method = getattr(self.db, create_method_name, None)
        if not create_method:
            results["errors"].append(
                f"DB has no method '{create_method_name}'"
            )
            return results

        try:
            wb = openpyxl.load_workbook(str(file_path))
        except Exception as e:
            results["errors"].append(f"Could not open file: {e}")
            return results

        ws = wb.active
        headers = [
            str(cell.value or "").strip().lower() for cell in ws[1]
        ]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                record = {
                    headers[i]: row[i]
                    for i in range(min(len(headers), len(row)))
                }
                record.pop("id", None)
                record = {k: v for k, v in record.items() if v is not None}
                if record:
                    create_method(record)
                    results["imported"] += 1
            except Exception as e:
                results["errors"].append(str(e))
                results["skipped"] += 1

        return results

    # ── Export entire knowledge base ──────────────────────────────────────────

    def export_all_knowledge(self, output_path) -> bool:
        """Export all knowledge tables as a multi-tab Excel workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Cover sheet
        cover = wb.active
        cover.title = "Knowledge Base"
        cover["A1"] = "ContractIQ Knowledge Base Export"
        cover["A1"].font = Font(bold=True, size=14, color="1a1f2e")
        cover["A2"] = f"Exported: {date.today().isoformat()}"
        cover["A4"] = "Contents:"
        cover["A4"].font = Font(bold=True)
        for i, table in enumerate(SUPPORTED_TABLES, 5):
            cover.cell(row=i, column=1, value=table)
        cover.column_dimensions["A"].width = 35

        header_fill = PatternFill(fill_type="solid", fgColor="1a1f2e")

        for table in SUPPORTED_TABLES:
            try:
                records = self._get_all_records(table)
                ws = wb.create_sheet(title=table[:31])

                if records:
                    headers = list(records[0].keys())
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    for r, rec in enumerate(records, 2):
                        for c, val in enumerate(rec.values(), 1):
                            ws.cell(row=r, column=c, value=val).alignment = Alignment(
                                vertical="top", wrap_text=True
                            )

                    # Auto-fit
                    for col in ws.columns:
                        max_len = max(
                            (len(str(cell.value or "")) for cell in col), default=10
                        )
                        ws.column_dimensions[col[0].column_letter].width = min(
                            max(max_len + 2, 12), 60
                        )
                    ws.freeze_panes = "A2"
                else:
                    ws.cell(row=1, column=1, value="(no records)")
            except Exception as e:
                print(f"  Warning: could not export {table}: {e}")

        wb.save(str(output_path))
        return True

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _get_all_records(self, table_name: str) -> list:
        get_method_name, _ = _TABLE_METHOD_MAP.get(table_name, (None, None))
        if not get_method_name:
            return []
        method = getattr(self.db, get_method_name, None)
        if not method:
            return []
        try:
            return method(active_only=False)
        except TypeError:
            return method()
