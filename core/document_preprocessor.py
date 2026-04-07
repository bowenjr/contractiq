"""
ContractIQ — Document Preprocessor (Stage 1 — Pure Python)
Zero LLM calls. Runs before the 7-pillar analysis engine.

Pipeline:
  1. _clean_raw_text      — remove noise (headers/footers, page numbers, TOC)
  2. _detect_sections     — locate structural headings via regex
  3. _build_markdown      — build structured markdown with ## / ### headers
  4. _build_section_index — keyword → section heading routing map

generate_tracker_sheet() produces the Excel contract-tracker workbook
and is unchanged from the previous version.
"""

import re
import json
from collections import Counter
from datetime import date as _date
from pathlib import Path
from typing import Callable, Dict, Any, List, Optional, Tuple

_POSITIONS_PATH = Path(__file__).parent.parent / "positions.json"

ProgressCB = Optional[Callable[[int, int, str, str, int], None]]

# item_type → pillar name (used by tracker sheet)
_TYPE_TO_PILLAR: Dict[str, str] = {
    "Payment":     "Money",
    "Liability":   "Risk & Liability",
    "Risk":        "Risk & Liability",
    "Deadline":    "Time",
    "Obligation":  "Administration",
    "Right":       "Relationships",
    "Condition":   "Scope",
    "Restriction": "Scope",
}

# Heading regex patterns — (compiled pattern, level)
# level 1 = top-level section, level 2 = sub-section
_HEADING_PATTERNS: List[Tuple[re.Pattern, int]] = [
    (re.compile(r"^(ARTICLE|Article)\s+\d+\b.*$",                           re.M), 1),
    (re.compile(r"^(SECTION|Section)\s+\d+\b.*$",                           re.M), 1),
    (re.compile(r"^\d+\.\s+[A-Z][A-Za-z\s]+",                               re.M), 1),
    (re.compile(r"^\d+\.\d+\s+[A-Z][A-Za-z\s]+",                            re.M), 2),
    (re.compile(r"^(SCHEDULE|Schedule|APPENDIX|Appendix|ANNEX|Annex)\s+[A-Z0-9]\b.*$",
                re.M), 1),
    (re.compile(r"^[A-Z][A-Z\s]{4,}$",                                      re.M), 1),
]

# Keywords scanned when building the section index for pillar routing
_INDEX_KEYWORDS: List[str] = [
    # Money / payment
    "payment", "invoice", "price", "cost", "fee", "retention", "holdback",
    "variation", "milestone", "liquidated", "damages", "escalation", "valuation",
    # Time
    "time", "schedule", "completion", "delay", "extension", "programme", "deadline",
    "commencement", "suspension",
    # Scope
    "scope", "work", "service", "supply", "design", "specification", "drawing",
    "exclusion", "provisional", "change",
    # Risk / liability
    "liability", "indemnity", "insurance", "warranty", "risk", "force majeure",
    "limitation", "consequential", "guarantee", "bond", "fitness",
    # Relationships
    "subcontract", "assignment", "novation", "personnel", "consent", "approval",
    "step-in", "flow-down", "back-to-back",
    # Administration
    "notice", "dispute", "arbitration", "governing", "jurisdiction",
    "confidential", "record", "audit", "instruction", "law",
    # Exit
    "termination", "terminate", "handover", "defect", "final account",
    "insolvency", "convenience", "remedy", "cure",
    # General
    "definition", "interpretation", "general",
]


def _load_positions() -> Dict:
    try:
        return json.loads(_POSITIONS_PATH.read_text())
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


class DocumentPreprocessor:

    def __init__(self, llm=None):
        """
        llm parameter accepted for backward compatibility but ignored.
        Stage 1 preprocessing is 100% pure Python — no LLM calls.
        """
        pass  # llm intentionally unused

    # ── Public entry point ────────────────────────────────────────────────────

    def preprocess(
        self,
        text: str,
        filename: str,
        doc_type: str = "General Contract",
        progress_callback: ProgressCB = None,
    ) -> Dict[str, Any]:
        """
        Pure-Python preprocessing pipeline.  No network calls.
        Completes in < 5 seconds on any document size.

        Returns:
          structured_markdown  — clean markdown with ## / ### section headers
          section_index        — {keyword: [heading, ...]} routing map
          sections             — [{heading, level, start_pos, end_pos, content}]
          section_count        — int
          word_count           — int
          operative_word_count — int (alias of word_count)
          noise_removed_pct    — float
          contractual_items    — [] (LLM flagging is Stage 2, done in analysis engine)
        """

        def _cb(message: str, percent: int) -> None:
            print(f"  Pre-processing: {message}")
            if progress_callback:
                progress_callback(0, 7, "Pre-processing", message, percent)

        original_len = len(text)

        # Step 1 — Clean
        _cb("Cleaning text...", 1)
        cleaned = self._clean_raw_text(text)
        noise_pct = round(
            max(0.0, (original_len - len(cleaned)) / max(1, original_len) * 100), 1
        )

        # Step 2 — Detect sections
        sections = self._detect_sections(cleaned)
        _cb(f"Detected {len(sections)} sections...", 2)

        # Attach content snippet to each section (used by index builder)
        for sec in sections:
            raw = cleaned[sec["start_pos"]: sec["end_pos"]]
            first_nl = raw.find("\n")
            sec["content"] = raw[first_nl + 1:].strip() if first_nl != -1 else ""

        # Step 3 — Build markdown
        _cb("Building structured markdown...", 3)
        word_count = len(cleaned.split())
        structured_md = self._build_markdown(cleaned, sections, filename, word_count)

        # Step 4 — Build section index
        _cb("Building section index...", 4)
        section_index = self._build_section_index(sections, cleaned)

        _cb(
            f"Complete — {len(sections)} sections, {noise_pct}% noise removed, "
            f"{word_count:,} words",
            5,
        )

        return {
            "structured_markdown":  structured_md,
            "section_index":        section_index,
            "sections":             sections,
            "section_count":        len(sections),
            "word_count":           word_count,
            "operative_word_count": word_count,
            "noise_removed_pct":    noise_pct,
            "contractual_items":    [],
        }

    def save_markdown(self, document_id: str, markdown: str, db) -> None:
        """Persist structured markdown to the documents table."""
        db.update_document(document_id, {"structured_markdown": markdown})

    # ── Step 1: Clean raw text ────────────────────────────────────────────────

    def _clean_raw_text(self, text: str) -> str:
        lines = text.splitlines()
        stripped_lines = [l.strip() for l in lines]

        # Detect repeated header/footer lines (> 3 occurrences)
        counts = Counter(
            l for l in stripped_lines
            if l and len(l) > 4 and not l.isdigit()
        )
        noise_lines = {line for line, cnt in counts.items() if cnt > 3}

        # Detect table of contents block (4+ consecutive dot-leader / number lines)
        toc_line_re = re.compile(r"^.{3,}\.{4,}\s*\d*\s*$|^\s*\d+\s*$")
        in_toc = False
        toc_start = toc_end = None
        run = 0
        for i, l in enumerate(stripped_lines):
            if toc_line_re.match(l):
                if not in_toc:
                    toc_start = i
                    in_toc = True
                run += 1
                toc_end = i
            else:
                if run >= 4:
                    break  # confirmed TOC block
                in_toc = False
                toc_start = None
                run = 0

        # Page number patterns:
        #   "Page 3", "Page 3 of 10", "- 3 -", "3", "5 | P a g e"
        page_re = re.compile(
            r"^\s*(?:page\s+\d+(?:\s+of\s+\d+)?|-\s*\d+\s*-|\d+\s*\|\s*P\s*a\s*g\s*e|\d+)\s*$",
            re.IGNORECASE,
        )

        cleaned: List[str] = []
        for i, (raw, stripped) in enumerate(zip(lines, stripped_lines)):
            if toc_start is not None and toc_start <= i <= toc_end:
                continue
            if stripped in noise_lines:
                continue
            if page_re.match(stripped):
                continue
            cleaned.append(raw.rstrip())

        result = "\n".join(cleaned)

        # Normalize whitespace
        result = re.sub(r"\r\n|\r", "\n", result)
        result = re.sub(r"\n{3,}", "\n\n", result)
        result = re.sub(r"[^\x09\x0A\x20-\x7E\u00A0-\uFFFF]", "", result)

        return result.strip()

    # ── Step 2: Detect sections ───────────────────────────────────────────────

    def _detect_sections(self, text: str) -> List[Dict[str, Any]]:
        """Return [{heading, level, start_pos, end_pos}] sorted by position."""
        hits: List[Tuple[int, str, int]] = []

        for pattern, level in _HEADING_PATTERNS:
            for m in pattern.finditer(text):
                heading = m.group(0).strip()
                if 3 < len(heading) < 120:
                    hits.append((m.start(), heading, level))

        # Deduplicate by start position (prefer lower level = more important)
        pos_map: Dict[int, Tuple[str, int]] = {}
        for pos, heading, level in hits:
            if pos not in pos_map or level < pos_map[pos][1]:
                pos_map[pos] = (heading, level)

        sorted_hits = sorted(pos_map.items())

        sections: List[Dict] = []
        for i, (start_pos, (heading, level)) in enumerate(sorted_hits):
            end_pos = (
                sorted_hits[i + 1][0] if i + 1 < len(sorted_hits) else len(text)
            )
            sections.append({
                "heading":   heading,
                "level":     level,
                "start_pos": start_pos,
                "end_pos":   end_pos,
            })

        return sections

    # ── Step 3: Build markdown ────────────────────────────────────────────────

    def _build_markdown(
        self,
        text: str,
        sections: List[Dict],
        filename: str = "",
        word_count: int = 0,
    ) -> str:
        """
        Convert cleaned text + detected sections into structured markdown.
        Metadata block at top; ## for level-1, ### for level-2 headings.
        All content is preserved verbatim — no summarisation.
        """
        parts: List[str] = []

        # Metadata header
        today = _date.today().isoformat()
        parts.append(f"# {filename}")
        parts.append(f"**Processed:** {today}  ")
        parts.append(f"**Sections detected:** {len(sections)}  ")
        parts.append(f"**Word count:** {word_count:,}  ")
        parts.append("---\n")

        if not sections:
            parts.append(_text_to_md_body(text))
            return "\n".join(parts)

        # Preamble — content before the first heading
        if sections[0]["start_pos"] > 0:
            preamble = text[: sections[0]["start_pos"]].strip()
            if preamble:
                parts.append("## Preamble\n")
                parts.append(_text_to_md_body(preamble))
                parts.append("")

        for sec in sections:
            heading = sec["heading"]
            level   = sec["level"]
            prefix  = "##" if level == 1 else "###"

            content = text[sec["start_pos"]: sec["end_pos"]].strip()
            first_nl = content.find("\n")
            body = content[first_nl:].strip() if first_nl != -1 else ""

            parts.append(f"{prefix} {heading}\n")
            if body:
                parts.append(_text_to_md_body(body))
            parts.append("")

        return "\n".join(parts)

    # ── Step 4: Build section index ───────────────────────────────────────────

    def _build_section_index(self, sections: List[Dict], text: str) -> Dict[str, List[str]]:
        """
        Map keywords → [section headings] for pillar routing in Stage 2.
        Checks both the heading text and the first 500 chars of section content.

        Example output:
          {"payment": ["Article 4 Payment Terms", "Appendix C Rates"],
           "termination": ["Article 8 Termination for Cause"]}
        """
        index: Dict[str, List[str]] = {}
        for section in sections:
            start = section["start_pos"]
            end   = section["end_pos"]
            # heading + first 500 chars of body
            snippet       = text[start: min(start + 500, end)].lower()
            heading_lower = section["heading"].lower()

            for kw in _INDEX_KEYWORDS:
                if kw in heading_lower or kw in snippet:
                    bucket = index.setdefault(kw, [])
                    if section["heading"] not in bucket:
                        bucket.append(section["heading"])

        return index

    # ── Tracker sheet ─────────────────────────────────────────────────────────

    def generate_tracker_sheet(
        self,
        document: Dict,
        contractual_items: List[Dict],
        analysis_results: Dict,
        output_path: Path,
    ) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.styles.protection import Protection
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        positions = _load_positions()
        std_pos = positions.get("standard_positions", {})

        wb = Workbook()
        wb.remove(wb.active)

        self._wb            = wb
        self._Font          = Font
        self._Fill          = PatternFill
        self._Align         = Alignment
        self._Border        = Border
        self._Side          = Side
        self._Protection    = Protection
        self._DataValidation = DataValidation
        self._gcl           = get_column_letter

        self._tab1_tracker(contractual_items, analysis_results, std_pos)
        self._tab2_obligations(contractual_items)
        self._tab3_legal(contractual_items)
        self._tab4_payment(contractual_items)
        self._tab5_resolved()

        wb.save(str(output_path))
        return output_path

    # ── Tracker Tab 1: Contract Tracker ───────────────────────────────────────

    def _tab1_tracker(
        self,
        items: List[Dict],
        analysis_results: Dict,
        std_pos: Dict,
    ) -> None:
        ws = self._wb.create_sheet("Contract Tracker")
        ws.sheet_properties.tabColor = "1a1f2e"
        ws.freeze_panes = "A2"

        HEADERS = [
            ("ID",                    12,  False),
            ("Article / Section",     28,  False),
            ("Item Type",             16,  False),
            ("Party",                 18,  False),
            ("Source Text",           45,  False),
            ("Plain English",         40,  False),
            ("Pillar",                18,  False),
            ("Severity",              12,  False),
            ("Our Standard Position", 30,  False),
            ("Deviation",             12,  False),
            ("Proposed Response",     30,  True),
            ("Redline Required",      16,  True),
            ("Internal Owner",        18,  True),
            ("Requires Legal",        14,  False),
            ("Lawyer Notes",          60,  True),
            ("Status",                14,  True),
            ("Date Resolved",         14,  True),
        ]

        NAVY   = "1a1f2e"
        BLUE   = "1d4ed8"
        WHITE  = "FFFFFF"
        BORDER = "d4cfc6"

        SEV_FILLS = {
            "Critical": "fee2e2",
            "High":     "fef3c7",
            "Medium":   "fefce8",
            "Low":      "f0fdf4",
        }

        def mkfill(hex_): return self._Fill(fill_type="solid", fgColor=hex_.lstrip("#"))
        def mkfont(bold=False, sz=10, color="1a1f2e"):
            return self._Font(name="Calibri", size=sz, bold=bold, color=color)
        def mkalign(h="left", v="top", wrap=True):
            return self._Align(horizontal=h, vertical=v, wrap_text=wrap)
        def mkborder():
            s = self._Side(style="thin", color=BORDER)
            return self._Border(left=s, right=s, top=s, bottom=s)

        for ci, (hdr, width, user_fill) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            hdr_colour = BLUE if user_fill else NAVY
            cell.font      = mkfont(bold=True, sz=10, color=WHITE)
            cell.fill      = mkfill(hdr_colour)
            cell.alignment = mkalign("center", "center")
            cell.border    = mkborder()
            ws.column_dimensions[self._gcl(ci)].width = width

        for ri, item in enumerate(items, 2):
            sev       = item.get("severity", "Low")
            cell_fill = mkfill(SEV_FILLS.get(sev, WHITE))
            itype     = item.get("item_type", "")
            pillar    = _TYPE_TO_PILLAR.get(itype, "Administration")
            req_legal = sev in ("Critical", "High")
            our_pos   = _match_position(itype, item.get("article", ""), std_pos)
            deviation = _check_deviation(item, our_pos)

            row_values = [
                item.get("item_id", f"C{ri-1:03d}"),
                item.get("article", ""),
                itype,
                item.get("party", ""),
                item.get("source_text", ""),
                item.get("plain_english", ""),
                pillar,
                sev,
                our_pos,
                deviation,
                "", "", "",
                "Yes" if req_legal else "No",
                "", "Open", "",
            ]

            for ci, (val, (_, _, user_fill)) in enumerate(zip(row_values, HEADERS), 1):
                cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
                cell.fill       = mkfill(WHITE) if user_fill else cell_fill
                cell.font       = mkfont(sz=9)
                cell.alignment  = mkalign()
                cell.border     = mkborder()
                cell.protection = self._Protection(locked=not user_fill)
            ws.row_dimensions[ri].height = 40

        ws.protection.sheet    = True
        ws.protection.password = "contractiq"
        ws.protection.enable()

        if len(items) > 0:
            dv = self._DataValidation(
                type="list",
                formula1='"Open,In Review,Agreed,Escalated"',
                allow_blank=True,
                showErrorMessage=False,
            )
            ws.add_data_validation(dv)
            dv.add(f"P2:P{len(items)+1}")

        ws.auto_filter.ref = f"A1:{self._gcl(len(HEADERS))}1"

    # ── Tracker Tab 2: Obligations ────────────────────────────────────────────

    def _tab2_obligations(self, items: List[Dict]) -> None:
        ws = self._wb.create_sheet("Obligations Register")
        ws.sheet_properties.tabColor = "1d4ed8"
        ws.freeze_panes = "A2"

        hdrs   = ["ID", "Party", "Description", "Trigger", "Deadline",
                  "Owner", "Status", "Notes"]
        widths = [10, 18, 50, 28, 18, 18, 14, 40]
        self._tracker_header(ws, hdrs, widths)

        row = 2
        for item in items:
            if item.get("item_type") not in ("Obligation", "Deadline", "Condition"):
                continue
            self._tracker_data_row(ws, row, [
                item.get("item_id", ""), item.get("party", ""),
                item.get("plain_english", ""),
                "", "", "", "Open", "",
            ], item.get("severity", "Low"))
            row += 1

        ws.auto_filter.ref = f"A1:H{row}"

    # ── Tracker Tab 3: Issues for Legal ──────────────────────────────────────

    def _tab3_legal(self, items: List[Dict]) -> None:
        ws = self._wb.create_sheet("Issues for Legal")
        ws.sheet_properties.tabColor = "dc2626"
        ws.freeze_panes = "A2"

        hdrs   = ["ID", "Article", "Source Text", "Risk Description",
                  "Severity", "Proposed Position", "Lawyer Notes", "Status"]
        widths = [10, 24, 45, 40, 12, 35, 50, 14]
        self._tracker_header(ws, hdrs, widths)

        row = 2
        for item in items:
            if item.get("severity") not in ("Critical", "High"):
                continue
            self._tracker_data_row(ws, row, [
                item.get("item_id", ""), item.get("article", ""),
                item.get("source_text", ""), item.get("plain_english", ""),
                item.get("severity", ""), "", "", "Open",
            ], item.get("severity", "Low"))
            row += 1

        ws.auto_filter.ref = f"A1:H{row}"

    # ── Tracker Tab 4: Payment & Commercial ───────────────────────────────────

    def _tab4_payment(self, items: List[Dict]) -> None:
        ws = self._wb.create_sheet("Payment & Commercial")
        ws.sheet_properties.tabColor = "16a34a"
        ws.freeze_panes = "A2"

        hdrs   = ["ID", "Article", "Source Text", "Risk",
                  "Our Position", "Response Required", "Status"]
        widths = [10, 24, 45, 35, 30, 20, 14]
        self._tracker_header(ws, hdrs, widths)

        row = 2
        for item in items:
            if item.get("item_type") not in ("Payment", "Liability", "Risk"):
                continue
            self._tracker_data_row(ws, row, [
                item.get("item_id", ""), item.get("article", ""),
                item.get("source_text", ""), item.get("plain_english", ""),
                "", "", "Open",
            ], item.get("severity", "Low"))
            row += 1

        ws.auto_filter.ref = f"A1:G{row}"

    # ── Tracker Tab 5: Resolved Items ─────────────────────────────────────────

    def _tab5_resolved(self) -> None:
        ws = self._wb.create_sheet("Resolved Items")
        ws.sheet_properties.tabColor = "6b7280"
        ws.freeze_panes = "A2"

        hdrs   = ["ID", "Article", "Item Type", "Severity",
                  "Plain English", "Resolution", "Date Resolved"]
        widths = [10, 24, 16, 12, 45, 40, 16]
        self._tracker_header(ws, hdrs, widths)

        note = ws.cell(
            row=2, column=1,
            value="Items will appear here when marked Agreed or Resolved in Contract Tracker.",
        )
        note.font = self._Font(name="Calibri", size=9, italic=True, color="6b7280")
        ws.merge_cells("A2:G2")

    # ── Shared helpers ────────────────────────────────────────────────────────

    def _tracker_header(self, ws, headers: List[str], widths: List[int]) -> None:
        NAVY = "1a1f2e"
        for ci, (hdr, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font      = self._Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill      = self._Fill(fill_type="solid", fgColor=NAVY)
            cell.alignment = self._Align(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = self._thin_border()
            ws.column_dimensions[self._gcl(ci)].width = w

    def _tracker_data_row(self, ws, row: int, values: List, severity: str = "Low") -> None:
        SEV_FILLS = {
            "Critical": "fee2e2", "High": "fef3c7",
            "Medium":   "fefce8", "Low":  "f0fdf4",
        }
        bg = SEV_FILLS.get(severity, "FFFFFF")
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=str(val) if val else "")
            cell.fill      = self._Fill(fill_type="solid", fgColor=bg)
            cell.font      = self._Font(name="Calibri", size=9, color="1a1f2e")
            cell.alignment = self._Align(horizontal="left", vertical="top", wrap_text=True)
            cell.border    = self._thin_border()
        ws.row_dimensions[row].height = 40

    def _thin_border(self):
        s = self._Side(style="thin", color="d4cfc6")
        return self._Border(left=s, right=s, top=s, bottom=s)


# ── Module-level helpers ──────────────────────────────────────────────────────

def _text_to_md_body(text: str) -> str:
    """Convert plain text block to clean markdown paragraphs (or tables)."""
    paras = re.split(r"\n{2,}", text.strip())
    out = []
    for para in paras:
        para = para.strip()
        if not para:
            continue
        lines = para.splitlines()
        if len(lines) >= 2 and any("|" in l or "\t" in l for l in lines):
            out.append(_lines_to_md_table(lines))
        else:
            out.append(para)
    return "\n\n".join(out)


def _lines_to_md_table(lines: List[str]) -> str:
    rows = []
    for line in lines:
        if "|" in line:
            cols = [c.strip() for c in line.split("|") if c.strip()]
        elif "\t" in line:
            cols = [c.strip() for c in line.split("\t") if c.strip()]
        else:
            cols = [line.strip()]
        rows.append(cols)

    if not rows:
        return "\n".join(lines)

    max_cols = max(len(r) for r in rows)
    for r in rows:
        while len(r) < max_cols:
            r.append("")

    header = "| " + " | ".join(rows[0]) + " |"
    sep    = "| " + " | ".join(["---"] * max_cols) + " |"
    data   = ["| " + " | ".join(r) + " |" for r in rows[1:]]
    return "\n".join([header, sep] + data)


def _match_position(item_type: str, article: str, std_pos: Dict) -> str:
    combined = (item_type + " " + article).lower()
    mapping = {
        "payment":   std_pos.get("payment_terms", ""),
        "retention": std_pos.get("retention", ""),
        "liability": std_pos.get("liability_cap", ""),
        "governing": std_pos.get("governing_law", ""),
        "dispute":   std_pos.get("dispute_resolution", ""),
        "notice":    std_pos.get("notice_method", ""),
        "eot":       std_pos.get("eot_notice_period", ""),
        "subcontract": std_pos.get("subcontracting", ""),
        "warranty":  std_pos.get("warranty_period", ""),
        "insurance": std_pos.get("insurance_gl", ""),
    }
    for kw, pos in mapping.items():
        if kw in combined and pos:
            return pos
    return ""


def _check_deviation(item: Dict, our_pos: str) -> str:
    if not our_pos:
        return "TBD"
    source = (item.get("source_text", "") + " " + item.get("plain_english", "")).lower()
    if "unlimited" in source and "limit" in our_pos.lower():
        return "Yes"
    if "60 days" in source and "30 days" in our_pos.lower():
        return "Yes"
    if "10%" in source and "5%" in our_pos.lower():
        return "Yes"
    return "TBD"
