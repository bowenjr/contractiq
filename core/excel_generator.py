"""
ContractIQ — Excel Workbook Generator
Produces a professional multi-tab Excel workbook from pillar analysis results.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List


def _safe(val, default="") -> str:
    if val is None or val == "" or str(val).lower() == "null":
        return default
    return str(val)


def _yn(val) -> str:
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, int):
        return "Yes" if val else "No"
    return str(val) if val is not None else ""


# ── Styling constants ─────────────────────────────────────────────────────────
HDR_FILL = "1a1f2e"       # dark navy
HDR_FONT = "FFFFFF"       # white
SUB_FILL = "c8a96e"       # gold
SUB_FONT = "1a1f2e"       # dark navy
PARCHMENT = "f8f6f0"
BORDER_CLR = "d4cfc6"

SEVERITY_FILLS = {
    "Critical": "fee2e2",
    "High":     "fef3c7",
    "Medium":   "fefce8",
    "Low":      "f0fdf4",
    "Info":     "f0f9ff",
}

STATUS_FILLS = {
    "Good":             "f0fdf4",
    "Issues Found":     "fef3c7",
    "Critical Issues":  "fee2e2",
    "Not Applicable":   "f3f4f6",
}

PILLAR_COLOURS = {
    "money":          "16a34a",
    "time":           "d97706",
    "scope":          "1d4ed8",
    "risk_liability": "dc2626",
    "relationships":  "7c3aed",
    "administration": "0891b2",
    "exit":           "6b7280",
}

TAB_COLOURS = {
    "Executive Summary":   "1a1f2e",
    "7 Pillars Analysis":  "c8a96e",
    "Negotiation Issues":  "dc2626",
    "Risk Register":       "d97706",
    "Obligations":         "1d4ed8",
    "Scope Gaps":          "7c3aed",
    "Action Tracker":      "16a34a",
    "Questions Answered":  "6b7280",
}


class ExcelGenerator:
    def __init__(self, reports_dir: Path):
        self.reports_dir = reports_dir
        self.reports_dir.mkdir(exist_ok=True)

    def generate(self, document: Dict, analysis: Dict, output_path: Path) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, GradientFill
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._wb = wb
        self._Font = Font
        self._Fill = PatternFill
        self._Align = Alignment
        self._Border = Border
        self._Side = Side
        self._get_col = get_column_letter

        pillars = analysis.get("pillars", [])
        recs = analysis.get("recommendations", {})
        risk_score = analysis.get("risk_score", {})
        dates = analysis.get("dates", {})

        self._tab_exec_summary(document, analysis, pillars, recs, risk_score)
        self._tab_pillars(pillars)
        self._tab_negotiation(pillars)
        self._tab_risk_register(pillars)
        self._tab_obligations(analysis.get("obligations", []))
        self._tab_scope_gaps(pillars)
        self._tab_action_tracker(recs)
        self._tab_questions(pillars)

        wb.save(str(output_path))
        return output_path

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _make_fill(self, hex_str: str):
        return self._Fill(fill_type="solid", fgColor=hex_str.lstrip("#"))

    def _hdr_font(self, bold=True, size=10, colour=HDR_FONT):
        return self._Font(name="Calibri", size=size, bold=bold, color=colour)

    def _body_font(self, bold=False, size=10, colour="1a1f2e"):
        return self._Font(name="Calibri", size=size, bold=bold, color=colour)

    def _thin_border(self):
        side = self._Side(style="thin", color=BORDER_CLR)
        return self._Border(left=side, right=side, top=side, bottom=side)

    def _wrap(self, horizontal="left", vertical="top", wrap=True):
        return self._Align(horizontal=horizontal, vertical=vertical,
                           wrap_text=wrap)

    def _header_row(self, ws, row_num: int, cols: List[str]):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=ci, value=col)
            cell.font = self._hdr_font()
            cell.fill = self._make_fill(HDR_FILL)
            cell.alignment = self._wrap("center", "middle")
            cell.border = self._thin_border()

    def _subheader_row(self, ws, row_num: int, cols: List[str]):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=ci, value=col)
            cell.font = self._hdr_font(colour=SUB_FONT)
            cell.fill = self._make_fill(SUB_FILL)
            cell.alignment = self._wrap("left", "middle")
            cell.border = self._thin_border()

    def _data_row(self, ws, row_num: int, values: List,
                  fill_hex: str = None, alt: bool = False):
        bg = fill_hex or (PARCHMENT if alt else "FFFFFF")
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=_safe(val))
            cell.font = self._body_font()
            cell.fill = self._make_fill(bg)
            cell.alignment = self._wrap()
            cell.border = self._thin_border()

    def _autofit(self, ws, min_w=12, max_w=60):
        for col in ws.columns:
            max_len = 0
            col_letter = self._get_col(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

    def _set_tab_colour(self, ws, tab_name: str):
        colour = TAB_COLOURS.get(tab_name, "1a1f2e")
        ws.sheet_properties.tabColor = colour.lstrip("#")

    def _add_filters(self, ws):
        from openpyxl.utils import get_column_letter
        max_col = ws.max_column
        if ws.max_row > 1 and max_col > 0:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(max_col)}{ws.max_row}"
            )

    # ── Tab 1: Executive Summary ──────────────────────────────────────────────

    def _tab_exec_summary(self, document, analysis, pillars, recs, risk_score):
        ws = self._wb.create_sheet("Executive Summary")
        self._set_tab_colour(ws, "Executive Summary")
        ws.freeze_panes = "A2"
        row = 1

        # Title
        ws.cell(row=row, column=1, value="ContractIQ — Executive Summary").font = \
            self._hdr_font(size=14)
        ws.cell(row=row, column=1).fill = self._make_fill(HDR_FILL)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        # Document info
        self._subheader_row(ws, row, ["Field", "Value", "", "Field", "Value", ""])
        row += 1
        meta = [
            ("Filename", _safe(document.get("filename"))),
            ("Document Type", _safe(analysis.get("doc_type"))),
            ("Upload Date", _safe(document.get("upload_date", "")[:10])),
            ("Analysis Date", datetime.now().strftime("%Y-%m-%d")),
            ("Contract Value", _safe(analysis.get("contract_value"))),
            ("Duration", _safe(analysis.get("contract_duration"))),
            ("Governing Law", _safe(analysis.get("governing_law"))),
            ("Counterparty", _safe(document.get("counterparty"))),
        ]
        for i in range(0, len(meta), 2):
            k1, v1 = meta[i]
            k2, v2 = meta[i+1] if i+1 < len(meta) else ("", "")
            self._data_row(ws, row, [k1, v1, "", k2, v2, ""], alt=(row % 2 == 0))
            row += 1

        row += 1

        # Risk score
        level = risk_score.get("level", "Unknown")
        score = risk_score.get("overall_score", 0)
        level_fill = {
            "Low": "f0fdf4", "Medium": "fef3c7",
            "High": "fee2e2", "Critical": "7f1d1d"
        }.get(level, "f3f4f6")

        ws.cell(row=row, column=1, value="RISK ASSESSMENT").font = \
            self._hdr_font(size=12)
        ws.cell(row=row, column=1).fill = self._make_fill(HDR_FILL)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        risk_rows = [
            ("Overall Risk Score", f"{score}/100"),
            ("Risk Level", level),
            ("Critical Flags", risk_score.get("critical_flags", 0)),
            ("High Flags", risk_score.get("high_flags", 0)),
            ("Score Rationale", risk_score.get("score_rationale", "")),
        ]
        for k, v in risk_rows:
            cell_k = ws.cell(row=row, column=1, value=k)
            cell_v = ws.cell(row=row, column=2, value=_safe(v))
            cell_k.font = self._body_font(bold=True)
            cell_v.font = self._body_font()
            if k == "Overall Risk Score":
                cell_v.fill = self._make_fill(level_fill)
                cell_v.font = self._hdr_font(colour="1a1f2e", size=12)
            ws.merge_cells(f"B{row}:F{row}")
            for c in [cell_k, cell_v]:
                c.border = self._thin_border()
                c.alignment = self._wrap()
            row += 1

        row += 1

        # Overall recommendation
        overall_rec = recs.get("overall_recommendation", "")
        rec_fill = {
            "Approve": "f0fdf4", "Approve with Amendments": "fef3c7",
            "Negotiate": "fef3c7", "Reject": "fee2e2",
        }.get(overall_rec, "f3f4f6")

        ws.cell(row=row, column=1, value="RECOMMENDATION").font = \
            self._hdr_font(size=12)
        ws.cell(row=row, column=1).fill = self._make_fill(HDR_FILL)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        cell_rec = ws.cell(row=row, column=1, value=overall_rec)
        cell_rec.font = self._hdr_font(colour="1a1f2e", size=12)
        cell_rec.fill = self._make_fill(rec_fill)
        ws.merge_cells(f"A{row}:F{row}")
        cell_rec.border = self._thin_border()
        row += 1

        rat = ws.cell(row=row, column=1,
                      value=recs.get("recommendation_rationale", ""))
        rat.font = self._body_font()
        rat.alignment = self._wrap()
        ws.merge_cells(f"A{row}:F{row}")
        rat.border = self._thin_border()
        row += 2

        # Key risks summary
        ws.cell(row=row, column=1, value="KEY RISKS SUMMARY").font = \
            self._hdr_font(size=11)
        ws.cell(row=row, column=1).fill = self._make_fill(SUB_FILL)
        ws.cell(row=row, column=1).font = self._hdr_font(colour=SUB_FONT, size=11)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1
        kr = ws.cell(row=row, column=1, value=recs.get("key_risks_summary", ""))
        kr.font = self._body_font()
        kr.alignment = self._wrap()
        ws.merge_cells(f"A{row}:F{row}")
        kr.border = self._thin_border()
        row += 2

        # Top 5 immediate actions
        actions = sorted(
            recs.get("immediate_actions", []),
            key=lambda x: int(x.get("priority", 5))
        )[:5]
        if actions:
            ws.cell(row=row, column=1, value="TOP IMMEDIATE ACTIONS").font = \
                self._hdr_font(size=11)
            ws.cell(row=row, column=1).fill = self._make_fill(HDR_FILL)
            ws.merge_cells(f"A{row}:F{row}")
            row += 1
            self._header_row(ws, row, ["#", "Action", "Reason", "Pillar", "Owner", ""])
            row += 1
            for i, a in enumerate(actions, 1):
                self._data_row(ws, row, [
                    i, a.get("action", ""), a.get("reason", ""),
                    a.get("pillar", ""), a.get("owner", ""), ""
                ], alt=(i % 2 == 0))
                row += 1

        self._autofit(ws)
        ws.row_dimensions[1].height = 28

    # ── Tab 2: 7 Pillars Analysis ─────────────────────────────────────────────

    def _tab_pillars(self, pillars: List[Dict]):
        ws = self._wb.create_sheet("7 Pillars Analysis")
        self._set_tab_colour(ws, "7 Pillars Analysis")
        ws.freeze_panes = "A2"
        row = 1

        cols = ["Pillar", "Status", "Score /100", "Summary",
                "Findings", "Red Flags", "Missing Protections"]
        self._header_row(ws, row, cols)
        row += 1

        for i, pr in enumerate(pillars, 1):
            status = pr.get("status", "Issues Found")
            bg = STATUS_FILLS.get(status, "FFFFFF")
            self._data_row(ws, row, [
                pr.get("pillar_name", ""),
                status,
                pr.get("score", ""),
                pr.get("summary", ""),
                len(pr.get("findings", [])),
                len(pr.get("red_flags", [])),
                len(pr.get("missing_protections", [])),
            ], fill_hex=bg)
            row += 1

        row += 1

        # Expanded findings per pillar
        for pr in pillars:
            pid = pr.get("pillar_id", "")
            colour = PILLAR_COLOURS.get(pid, "1a1f2e")

            ws.cell(row=row, column=1,
                    value=f"{pr.get('pillar_name', '')} — Detailed Findings")
            ws.cell(row=row, column=1).font = self._hdr_font(colour="FFFFFF", size=11)
            ws.cell(row=row, column=1).fill = self._make_fill(colour)
            ws.merge_cells(f"A{row}:G{row}")
            row += 1

            findings = pr.get("findings", [])
            if findings:
                self._subheader_row(ws, row,
                    ["Finding", "Detail", "Severity", "Clause Ref",
                     "Source Excerpt", "Deviation", "Action"])
                row += 1
                for j, f in enumerate(findings):
                    sev = f.get("severity", "")
                    bg = SEVERITY_FILLS.get(sev, "FFFFFF")
                    self._data_row(ws, row, [
                        f.get("finding", ""),
                        f.get("detail", ""),
                        sev,
                        f.get("clause_reference", ""),
                        f.get("source_excerpt", ""),
                        f.get("deviation_from_position", ""),
                        f.get("recommended_action", ""),
                    ], fill_hex=bg)
                    row += 1

            missing = pr.get("missing_protections", [])
            if missing:
                ws.cell(row=row, column=1, value="Missing Protections:").font = \
                    self._body_font(bold=True)
                ws.merge_cells(f"A{row}:G{row}")
                row += 1
                for m in missing:
                    ws.cell(row=row, column=1, value=f"  ⚠  {m}")
                    ws.cell(row=row, column=1).font = self._body_font(colour="dc2626")
                    ws.merge_cells(f"A{row}:G{row}")
                    row += 1

            row += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 3: Negotiation Issues ─────────────────────────────────────────────

    def _tab_negotiation(self, pillars: List[Dict]):
        ws = self._wb.create_sheet("Negotiation Issues")
        self._set_tab_colour(ws, "Negotiation Issues")
        ws.freeze_panes = "A2"

        cols = ["ID", "Pillar", "Issue", "Current Position",
                "Primary Ask", "Fallback", "Priority",
                "Requires Legal", "Status"]
        self._header_row(ws, 1, cols)
        row = 2
        idx = 1
        for pr in pillars:
            pillar_name = pr.get("pillar_name", "")
            for np in pr.get("negotiation_points", []):
                pri = np.get("priority", "Medium")
                bg = {"High": "fee2e2", "Medium": "fef3c7", "Low": "f0fdf4"}.get(
                    pri, "FFFFFF")
                self._data_row(ws, row, [
                    idx, pillar_name,
                    np.get("issue", ""),
                    np.get("current_position", ""),
                    np.get("primary_ask", ""),
                    np.get("fallback", ""),
                    pri,
                    _yn(np.get("requires_legal", False)),
                    "Open",
                ], fill_hex=bg)
                row += 1
                idx += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 4: Risk Register ──────────────────────────────────────────────────

    def _tab_risk_register(self, pillars: List[Dict]):
        ws = self._wb.create_sheet("Risk Register")
        self._set_tab_colour(ws, "Risk Register")
        ws.freeze_panes = "A2"

        cols = ["ID", "Pillar", "Flag", "Severity",
                "Description", "Location", "Source Excerpt", "Mitigation"]
        self._header_row(ws, 1, cols)
        row = 2
        idx = 1

        # Sort: critical first
        all_flags = []
        for pr in pillars:
            for flag in pr.get("red_flags", []):
                all_flags.append((pr.get("pillar_name", ""), flag))
        sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
        all_flags.sort(key=lambda x: sev_order.get(x[1].get("severity", ""), 9))

        for pillar_name, flag in all_flags:
            sev = flag.get("severity", "")
            bg = SEVERITY_FILLS.get(sev, "FFFFFF")
            self._data_row(ws, row, [
                idx, pillar_name,
                flag.get("flag", ""),
                sev,
                flag.get("description", ""),
                flag.get("location", ""),
                flag.get("source_excerpt", ""),
                "",  # mitigation — for user to fill
            ], fill_hex=bg)
            row += 1
            idx += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 5: Obligations ────────────────────────────────────────────────────

    def _tab_obligations(self, obligations: List[Dict]):
        ws = self._wb.create_sheet("Obligations")
        self._set_tab_colour(ws, "Obligations")
        ws.freeze_panes = "A2"

        cols = ["ID", "Party", "Type", "Description",
                "Trigger", "Deadline", "Notice Required",
                "Internal Owner", "Status"]
        self._header_row(ws, 1, cols)
        row = 2
        for i, ob in enumerate(obligations, 1):
            self._data_row(ws, row, [
                i,
                ob.get("party", ""),
                ob.get("obligation_type", ""),
                ob.get("description", ""),
                ob.get("trigger", ""),
                ob.get("deadline", ""),
                ob.get("notice_required", ""),
                ob.get("owner", ""),
                ob.get("status", "Open"),
            ], alt=(i % 2 == 0))
            row += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 6: Scope Gaps ─────────────────────────────────────────────────────

    def _tab_scope_gaps(self, pillars: List[Dict]):
        ws = self._wb.create_sheet("Scope Gaps")
        self._set_tab_colour(ws, "Scope Gaps")
        ws.freeze_panes = "A2"

        cols = ["ID", "Requirement", "Source Document",
                "In Our Quote", "Priced", "Owner",
                "Gap Status", "Comments"]
        self._header_row(ws, 1, cols)
        row = 2
        idx = 1

        # Pull scope findings from SCOPE pillar
        scope_pillar = next(
            (p for p in pillars if p.get("pillar_id") == "scope"), None
        )
        if scope_pillar:
            for f in scope_pillar.get("findings", []):
                self._data_row(ws, row, [
                    idx,
                    f.get("finding", ""),
                    f.get("clause_reference", ""),
                    "",  # user fills
                    "",  # user fills
                    "",  # user fills
                    "Review",
                    f.get("detail", ""),
                ], alt=(idx % 2 == 0))
                row += 1
                idx += 1

            for m in scope_pillar.get("missing_protections", []):
                self._data_row(ws, row, [
                    idx, m, "Missing", "No", "No", "", "Gap", ""
                ], fill_hex="fee2e2")
                row += 1
                idx += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 7: Action Tracker ─────────────────────────────────────────────────

    def _tab_action_tracker(self, recs: Dict):
        ws = self._wb.create_sheet("Action Tracker")
        self._set_tab_colour(ws, "Action Tracker")
        ws.freeze_panes = "A2"

        cols = ["Priority", "Action", "Pillar", "Owner",
                "Due Date", "Status", "Notes"]
        self._header_row(ws, 1, cols)
        row = 2

        actions = sorted(
            recs.get("immediate_actions", []),
            key=lambda x: int(x.get("priority", 5))
        )
        for i, a in enumerate(actions, 1):
            self._data_row(ws, row, [
                a.get("priority", i),
                a.get("action", ""),
                a.get("pillar", ""),
                a.get("owner", ""),
                "",  # due date — user fills
                "Open",
                "",  # notes — user fills
            ], alt=(i % 2 == 0))
            row += 1

        # Also add legal / estimating / operations items
        for section, owner in [
            ("legal_escalation_items", "Legal"),
            ("estimating_review_items", "Estimating"),
            ("operations_review_items", "Operations"),
        ]:
            for item in recs.get(section, []):
                self._data_row(ws, row, [
                    "", item, "", owner, "", "Open", ""
                ], alt=(row % 2 == 0))
                row += 1

        self._autofit(ws)
        self._add_filters(ws)

    # ── Tab 8: Questions Answered ─────────────────────────────────────────────

    def _tab_questions(self, pillars: List[Dict]):
        ws = self._wb.create_sheet("Questions Answered")
        self._set_tab_colour(ws, "Questions Answered")
        ws.freeze_panes = "A2"

        cols = ["Pillar", "Question", "Answer", "Found in Document"]
        self._header_row(ws, 1, cols)
        row = 2

        for pr in pillars:
            pillar_name = pr.get("pillar_name", "")
            qa = pr.get("questions_answered", {})
            if not qa:
                continue
            for question, answer in qa.items():
                found = "No" if str(answer).strip() in (
                    "Not Found", "Not Applicable", "N/A", ""
                ) else "Yes"
                bg = "f0fdf4" if found == "Yes" else "fef3c7"
                self._data_row(ws, row, [
                    pillar_name, question, _safe(answer), found
                ], fill_hex=bg)
                row += 1

        self._autofit(ws)
        self._add_filters(ws)
